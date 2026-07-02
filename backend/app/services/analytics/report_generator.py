"""Report generation service — daily, weekly, monthly, executive reports.

Generates report data objects and optionally exports them to PDF (via
ReportLab) and Excel (via openpyxl).  Both libraries are lightweight and
require no external services.

Install (add to requirements.txt):
    reportlab==4.2.2
    openpyxl==3.1.3
"""
from __future__ import annotations
import io
import logging
from datetime import datetime, timezone, timedelta, date

from app.extensions import db
from app.models.report import Report
from app.services.analytics.reach_analytics import get_reach_summary, get_reach_trend
from app.services.analytics.engagement_analytics import get_engagement_summary, get_platform_engagement
from app.services.analytics.growth_analytics import get_growth_metrics
from app.services.analytics.campaign_analytics import get_campaign_summary

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------

def _today_utc() -> date:
    return datetime.now(timezone.utc).date()


def _daily_window() -> tuple[datetime, datetime]:
    today = _today_utc()
    start = datetime(today.year, today.month, today.day, 0,  0,  0, tzinfo=timezone.utc)
    end   = datetime(today.year, today.month, today.day, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def _weekly_window() -> tuple[datetime, datetime]:
    today   = _today_utc()
    monday  = today - timedelta(days=today.weekday())
    sunday  = monday + timedelta(days=6)
    start   = datetime(monday.year, monday.month, monday.day,  0,  0,  0, tzinfo=timezone.utc)
    end     = datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def _monthly_window() -> tuple[datetime, datetime]:
    today  = _today_utc()
    start  = datetime(today.year, today.month, 1,  0,  0,  0, tzinfo=timezone.utc)
    import calendar
    last_day = calendar.monthrange(today.year, today.month)[1]
    end    = datetime(today.year, today.month, last_day, 23, 59, 59, tzinfo=timezone.utc)
    return start, end


def _custom_window(period_start: str, period_end: str) -> tuple[datetime, datetime]:
    fmt = '%Y-%m-%d'
    s = datetime.strptime(period_start, fmt).replace(tzinfo=timezone.utc)
    e = datetime.strptime(period_end, fmt).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    return s, e


# ---------------------------------------------------------------------------
# Data assembly
# ---------------------------------------------------------------------------

def _build_report_data(district_id: str, start: datetime, end: datetime,
                       report_type: str) -> dict:
    """Assemble all analytics sections into a single report dict."""
    data = {
        'report_type': report_type,
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'period': {'start': start.isoformat(), 'end': end.isoformat()},
        'reach': get_reach_summary(district_id, start, end),
        'engagement': get_engagement_summary(district_id, start, end),
        'growth': get_growth_metrics(district_id, start, end),
        'campaigns': get_campaign_summary(district_id, start, end),
        'platform_engagement': get_platform_engagement(district_id, start, end),
    }
    if report_type in ('weekly', 'monthly', 'executive'):
        data['reach_trend'] = get_reach_trend(
            district_id,
            days=7 if report_type == 'weekly' else 30,
        )
    return data


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_report(
    district_id: str,
    report_type: str,
    generated_by: str | None = None,
    period_start: str | None = None,
    period_end: str | None = None,
) -> Report:
    """Create and persist a report record with assembled data.

    Args:
        district_id:   Tenant scope.
        report_type:   ``'daily'`` | ``'weekly'`` | ``'monthly'`` | ``'executive'`` | ``'custom'``
        generated_by:  User ID requesting the report.
        period_start:  ISO date string (required for ``'custom'``).
        period_end:    ISO date string (required for ``'custom'``).

    Returns:
        Persisted :class:`Report` model instance (status=``'ready'``).
    """
    if report_type == 'daily':
        start, end = _daily_window()
    elif report_type == 'weekly':
        start, end = _weekly_window()
    elif report_type == 'monthly':
        start, end = _monthly_window()
    elif report_type in ('executive', 'custom'):
        if not period_start or not period_end:
            raise ValueError('period_start and period_end are required for custom/executive reports.')
        start, end = _custom_window(period_start, period_end)
    else:
        raise ValueError(f"Unknown report_type '{report_type}'. Use: daily/weekly/monthly/executive/custom.")

    titles = {
        'daily': 'Daily Social Media Report',
        'weekly': 'Weekly Social Media Report',
        'monthly': 'Monthly Social Media Report',
        'executive': 'Executive Summary Report',
        'custom': 'Custom Report',
    }

    report = Report(
        district_id=district_id,
        report_type=report_type,
        title=titles.get(report_type, 'Report'),
        period_start=start.date().isoformat(),
        period_end=end.date().isoformat(),
        generated_by=generated_by,
        status='pending',
    )
    db.session.add(report)
    db.session.flush()

    try:
        report.data = _build_report_data(district_id, start, end, report_type)
        report.status = 'ready'
        report.generated_at = datetime.now(timezone.utc).isoformat()
    except Exception as exc:
        logger.exception('Report generation failed for district %s', district_id)
        report.status = 'failed'
        report.error_message = str(exc)

    db.session.commit()
    return report


# ---------------------------------------------------------------------------
# PDF Export
# ---------------------------------------------------------------------------

def export_pdf(report: Report) -> bytes:
    """Generate a PDF byte stream for the given report.

    Uses ReportLab's Platypus layout engine.  Falls back to a plain text
    representation if ReportLab is not installed.

    Returns:
        Raw PDF bytes.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story  = []

        story.append(Paragraph(report.title, styles['Title']))
        story.append(Paragraph(
            f"Period: {report.period_start} → {report.period_end} | "
            f"Generated: {report.generated_at}",
            styles['Normal'],
        ))
        story.append(Spacer(1, 12))

        # --- Reach section
        reach = report.data.get('reach', {})
        story.append(Paragraph('Reach Summary', styles['Heading2']))
        reach_data = [
            ['Metric', 'Value'],
            ['Posts Published', reach.get('total_posts_published', 0)],
            ['Total Impressions', reach.get('total_impressions', 0)],
            ['Collected Posts', reach.get('total_collected_posts', 0)],
        ]
        t = Table(reach_data, hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

        # --- Engagement section
        eng = report.data.get('engagement', {}).get('outbound', {})
        story.append(Paragraph('Engagement Summary', styles['Heading2']))
        eng_data = [
            ['Metric', 'Value'],
            ['Total Likes',      eng.get('likes', 0)],
            ['Total Comments',   eng.get('comments', 0)],
            ['Total Shares',     eng.get('shares', 0)],
            ['Engagement Rate',  f"{eng.get('engagement_rate_pct', 0)}%"],
        ]
        t2 = Table(eng_data, hAlign='LEFT')
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ]))
        story.append(t2)
        story.append(Spacer(1, 12))

        # --- AI Flags
        inbound = report.data.get('engagement', {}).get('inbound', {})
        story.append(Paragraph('AI Intelligence Summary', styles['Heading2']))
        ai_data = [
            ['Category', 'Count', 'Rate'],
            ['Complaints', inbound.get('complaints', 0), f"{inbound.get('complaint_rate_pct', 0)}%"],
            ['Emergencies', inbound.get('emergencies', 0), f"{inbound.get('emergency_rate_pct', 0)}%"],
            ['Spam', inbound.get('spam', 0), '-'],
        ]
        t3 = Table(ai_data, hAlign='LEFT')
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ]))
        story.append(t3)

        doc.build(story)
        return buffer.getvalue()

    except ImportError:
        logger.warning('reportlab not installed — returning plain text PDF fallback')
        return _plain_text_fallback(report).encode('utf-8')


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_excel(report: Report) -> bytes:
    """Generate an Excel (.xlsx) byte stream for the given report.

    Uses openpyxl.  Falls back to CSV if not installed.

    Returns:
        Raw .xlsx bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = 'Overview'

        header_font  = Font(bold=True, color='FFFFFF')
        header_fill  = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

        def _header_row(ws_sheet, row_data, row_num):
            for col, val in enumerate(row_data, 1):
                cell = ws_sheet.cell(row=row_num, column=col, value=val)
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = Alignment(horizontal='center')

        # --- Overview sheet
        ws.append(['Report', report.title])
        ws.append(['Period', f"{report.period_start} to {report.period_end}"])
        ws.append(['Generated', report.generated_at or ''])
        ws.append([])

        reach = report.data.get('reach', {})
        ws.append(['REACH SUMMARY'])
        _header_row(ws, ['Metric', 'Value'], ws.max_row)
        metrics = [
            ('Posts Published',  reach.get('total_posts_published', 0)),
            ('Total Impressions', reach.get('total_impressions', 0)),
            ('Collected Posts',   reach.get('total_collected_posts', 0)),
        ]
        for m in metrics:
            ws.append(list(m))

        ws.append([])
        eng = report.data.get('engagement', {}).get('outbound', {})
        ws.append(['ENGAGEMENT SUMMARY'])
        _header_row(ws, ['Metric', 'Value'], ws.max_row)
        eng_metrics = [
            ('Likes',           eng.get('likes', 0)),
            ('Comments',        eng.get('comments', 0)),
            ('Shares',          eng.get('shares', 0)),
            ('Views',           eng.get('views', 0)),
            ('Engagement Rate', f"{eng.get('engagement_rate_pct', 0)}%"),
        ]
        for m in eng_metrics:
            ws.append(list(m))

        # --- Platform breakdown sheet
        ws2 = wb.create_sheet('Platform Breakdown')
        ws2.append(['Platform', 'Posts', 'Likes', 'Comments', 'Shares', 'Total Engagement'])
        for row in report.data.get('platform_engagement', []):
            ws2.append([
                row.get('platform', ''),
                row.get('posts', 0),
                row.get('likes', 0),
                row.get('comments', 0),
                row.get('shares', 0),
                row.get('total_engagement', 0),
            ])

        # --- Campaigns sheet
        ws3 = wb.create_sheet('Campaigns')
        ws3.append(['Campaign', 'Posts', 'Likes', 'Comments', 'Shares', 'Views', 'Engagement Rate'])
        for row in report.data.get('campaigns', []):
            ws3.append([
                row.get('campaign', ''),
                row.get('posts', 0),
                row.get('likes', 0),
                row.get('comments', 0),
                row.get('shares', 0),
                row.get('views', 0),
                f"{row.get('engagement_rate_pct', 0)}%",
            ])

        # Auto-size columns
        for ws_sheet in [ws, ws2, ws3]:
            for col in ws_sheet.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.warning('openpyxl not installed — returning CSV fallback')
        return _csv_fallback(report).encode('utf-8')


# ---------------------------------------------------------------------------
# Fallback helpers
# ---------------------------------------------------------------------------

def _plain_text_fallback(report: Report) -> str:
    lines = [
        report.title,
        f"Period: {report.period_start} to {report.period_end}",
        '',
        str(report.data),
    ]
    return '\n'.join(lines)


def _csv_fallback(report: Report) -> str:
    import csv, io as _io
    buf = _io.StringIO()
    w = csv.writer(buf)
    w.writerow(['metric', 'value'])
    reach = report.data.get('reach', {})
    w.writerow(['posts_published', reach.get('total_posts_published', 0)])
    w.writerow(['total_impressions', reach.get('total_impressions', 0)])
    return buf.getvalue()
