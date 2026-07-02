"""Named social report exports for Phase 9."""
from __future__ import annotations

import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.collected_post import CollectedPost
from app.models.social_comment import SocialComment
from app.services.social.collector_dashboard_service import COMMENT_LIKE_TYPES, get_collector_dashboard


COMMENTS_FILENAME = 'comments_report.xlsx'
COMPLAINTS_FILENAME = 'complaints_report.xlsx'
MONTHLY_SOCIAL_FILENAME = 'monthly_social_report.pdf'


def _month_window() -> tuple[datetime, int]:
    today = datetime.now(timezone.utc).date()
    start = datetime(today.year, today.month, 1, tzinfo=timezone.utc)
    return start, today.day


def _format_dt(value) -> str:
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value)


def _autosize(workbook) -> None:
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_len = max((len(str(cell.value or '')) for cell in column), default=10)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 4, 48)


def _style_header(sheet, row: int = 1) -> None:
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    for cell in sheet[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


def _comment_rows(district_id: str, complaints_only: bool = False) -> list[list]:
    rows = []

    comment_query = SocialComment.query.filter_by(district_id=district_id)
    collected_query = CollectedPost.query.filter(
        CollectedPost.district_id == district_id,
        CollectedPost.content_type.in_(COMMENT_LIKE_TYPES),
    )
    if complaints_only:
        comment_query = comment_query.filter(SocialComment.is_complaint.is_(True))
        collected_query = collected_query.filter(CollectedPost.is_complaint.is_(True))

    for comment in comment_query.order_by(SocialComment.created_at.desc()).all():
        department = 'Unassigned'
        issue_type = ''
        if comment.analysis:
            issue_type = comment.analysis.issue_type or ''
            if comment.analysis.service_request and comment.analysis.service_request.department:
                department = comment.analysis.service_request.department.name

        rows.append([
            comment.id,
            'owned_post_comment',
            comment.platform,
            comment.author_name or comment.author_username or '',
            comment.text,
            comment.sentiment or '',
            comment.sentiment_score or '',
            'yes' if comment.is_complaint else 'no',
            'yes' if comment.is_emergency else 'no',
            'yes' if comment.is_spam else 'no',
            issue_type,
            department,
            comment.ai_status,
            _format_dt(comment.created_at),
        ])

    for item in collected_query.order_by(CollectedPost.created_at.desc()).all():
        rows.append([
            item.id,
            item.content_type,
            item.platform,
            item.author_username or item.author_platform_id or '',
            item.raw_text,
            item.sentiment or '',
            item.sentiment_score or '',
            'yes' if item.is_complaint else 'no',
            'yes' if item.is_emergency else 'no',
            'yes' if item.is_spam else 'no',
            '',
            'Unassigned',
            item.ai_status,
            _format_dt(item.created_at),
        ])

    return rows


def export_comments_excel(district_id: str) -> bytes:
    return _export_comments_workbook(district_id, complaints_only=False)


def export_complaints_excel(district_id: str) -> bytes:
    return _export_comments_workbook(district_id, complaints_only=True)


def _export_comments_workbook(district_id: str, complaints_only: bool) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Complaints' if complaints_only else 'Comments'

    sheet.append([
        'ID',
        'Source Type',
        'Platform',
        'Author',
        'Text',
        'Sentiment',
        'Sentiment Score',
        'Complaint',
        'Emergency',
        'Spam',
        'Issue Type',
        'Department',
        'AI Status',
        'Created At',
    ])
    _style_header(sheet)

    for row in _comment_rows(district_id, complaints_only=complaints_only):
        sheet.append(row)

    if sheet.max_row == 1:
        sheet.append(['No data available'])

    _autosize(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_monthly_social_pdf(district_id: str) -> bytes:
    start, days = _month_window()
    dashboard = get_collector_dashboard(district_id, days=days)
    widgets = dashboard['widgets']
    charts = dashboard['charts']

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title='Monthly Social Report')
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Monthly Social Report', styles['Title']),
        Paragraph(f"Period: {start.date().isoformat()} to {datetime.now(timezone.utc).date().isoformat()}", styles['Normal']),
        Spacer(1, 12),
    ]

    story.append(Paragraph('Collector Summary', styles['Heading2']))
    story.append(_pdf_table([
        ['Metric', 'Value'],
        ['Total Comments', widgets['total_comments']],
        ['Positive', widgets['positive']],
        ['Negative', widgets['negative']],
        ['Complaints', widgets['complaints']],
    ], colors.HexColor('#1D4ED8')))
    story.append(Spacer(1, 12))

    sentiment_totals = {'positive': 0, 'negative': 0, 'neutral': 0, 'mixed': 0}
    for row in charts['sentiment_trend']:
        for key in sentiment_totals:
            sentiment_totals[key] += row.get(key, 0)

    story.append(Paragraph('Sentiment Trend', styles['Heading2']))
    story.append(_pdf_table([
        ['Sentiment', 'Count'],
        ['Positive', sentiment_totals['positive']],
        ['Negative', sentiment_totals['negative']],
        ['Neutral', sentiment_totals['neutral']],
        ['Mixed', sentiment_totals['mixed']],
    ], colors.HexColor('#16A34A')))
    story.append(Spacer(1, 12))

    department_rows = [['Department', 'Complaints']]
    department_rows.extend([
        [row['department'], row['complaints']]
        for row in charts['department_trend'][:10]
    ])
    story.append(Paragraph('Department Trend', styles['Heading2']))
    story.append(_pdf_table(department_rows, colors.HexColor('#F59E0B')))

    doc.build(story)
    return buffer.getvalue()


def _pdf_table(data: list[list], header_color) -> Table:
    table = Table(data, hAlign='LEFT', repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))
    return table
